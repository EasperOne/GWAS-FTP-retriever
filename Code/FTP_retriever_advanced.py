import pandas as pd
import openpyxl
from openpyxl.styles import Font
import colorsys
import pickle
import logging
import os
import json
import ieugwaspy as ieu

# Set up logging
logging.basicConfig(filename='ieugwas_process.log', level=logging.INFO,
                    format='%(asctime)s - %(levelname)s - %(message)s')

# Define the output Excel file path
OUTPUT_FILE = "ieugwas_tnf_output_1.xlsx"

def generate_color_palette(n):
    colors = []
    for i in range(n):
        hue = i / n
        rgb = colorsys.hsv_to_rgb(hue, 0.8, 0.8)
        colors.append('FF{:02X}{:02X}{:02X}'.format(int(rgb[0]*255), int(rgb[1]*255), int(rgb[2]*255)))
    return colors

def color_code_ids(df, existing_colors=None):
    if existing_colors is None:
        existing_colors = {}
    
    unique_ids = df['id'].dropna().unique()
    new_ids = [id for id in unique_ids if id not in existing_colors]
    
    if new_ids:
        new_colors = generate_color_palette(len(new_ids))
        existing_colors.update(dict(zip(new_ids, new_colors)))
    
    df['id_color'] = df['id'].map(existing_colors)
    
    return df, existing_colors

def search_and_process_data():
    logging.info("Starting IEUGWAS data processing for 'tumor necrosis factor'")
    
    try:
        # Load JWT token from .ieugwaspy.json file
        jwt_path = os.path.expanduser("~/.ieugwaspy.json")
        if os.path.exists(jwt_path):
            with open(jwt_path, 'r') as f:
                jwt_token = json.load(f).get("jwt")
                os.environ['IEUGWASPY_JWT'] = jwt_token
        else:
            logging.error("JWT token file not found.")
            print("JWT token file not found.")
            return

        # Query the GWAS database
        gwas_info = ieu.gwasinfo()

        # Print and log the structure of gwas_info
        print("Type of gwas_info:", type(gwas_info))
        logging.info(f"Type of gwas_info: {type(gwas_info)}")

        # Check if gwas_info is a dictionary and convert to a list
        if isinstance(gwas_info, dict):
            gwas_entries = list(gwas_info.values())
            print("First 5 entries of gwas_entries:", gwas_entries[:5])
            logging.info(f"First 5 entries of gwas_entries: {gwas_entries[:5]}")

            # Filter for studies related to tumor necrosis factor
            search_results = [study for study in gwas_entries if isinstance(study, dict) and "TNF" in study.get('trait', '').lower()]
        else:
            logging.error("GWAS Info is not in the expected dict format.")
            return

        logging.info(f"Found {len(search_results)} studies related to 'tumor necrosis factor'")
        
        # Convert search results to DataFrame
        df = pd.DataFrame(search_results)
        
        # Select relevant columns
        selected_columns = ['id', 'trait', 'year', 'author', 'pmid', 'population', 
                            'sample_size', 'nsnp', 'unit', 'sex', 'category', 'subcategory']
        df = df[selected_columns]
        
        # Check if the output Excel file exists
        try:
            existing_data = pd.read_excel(OUTPUT_FILE, sheet_name="IEUGWAS TNF Data", dtype=str, engine='openpyxl')
            with open(OUTPUT_FILE.replace('.xlsx', '_color_map.pkl'), 'rb') as f:
                color_map = pickle.load(f)
            
            # Identify new studies
            existing_ids = set(existing_data['id'])
            new_ids = set(df['id']) - existing_ids
            
            # Filter df to include only new studies
            df = df[df['id'].isin(new_ids)]
            
            # Append new data to existing data
            combined_data = pd.concat([existing_data, df], ignore_index=True)
            logging.info(f"Appending {len(df)} new studies to existing data")
        except FileNotFoundError:
            combined_data = df
            color_map = {}
            logging.info("Creating new output file")
        except Exception as e:
            logging.error(f"Error reading existing file: {e}")
            logging.info("Creating new output file")
            combined_data = df
            color_map = {}
        
        # Color-code study IDs
        result_df, updated_color_map = color_code_ids(combined_data, color_map)
        
        # Create a new workbook or load existing one
        if os.path.exists(OUTPUT_FILE):
            wb = openpyxl.load_workbook(OUTPUT_FILE)
            ws = wb["IEUGWAS TNF Data"]
            # Clear existing data (except header)
            ws.delete_rows(2, ws.max_row)
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "IEUGWAS TNF Data"
        
        # Write data to the worksheet
        for r, row in enumerate(result_df.values, start=1):
            for c, value in enumerate(row, start=1):
                ws.cell(row=r, column=c, value=value)
        
        # Apply color formatting to ID column
        id_col = result_df.columns.get_loc("id") + 1
        for r, color in enumerate(result_df['id_color'], start=2):
            if pd.notna(color):
                cell = ws.cell(row=r, column=id_col)
                cell.font = Font(color=color)
        
        # Save the workbook
        wb.save(OUTPUT_FILE)
        
        # Save the updated color map
        with open(OUTPUT_FILE.replace('.xlsx', '_color_map.pkl'), 'wb') as f:
            pickle.dump(updated_color_map, f)
        
        logging.info(f"Data processed and exported to {OUTPUT_FILE}")
        logging.info(f"Total studies in output: {len(result_df)}")
        
    except Exception as e:
        logging.error(f"An error occurred during processing: {e}")
        print(f"An error occurred: {e}")

# Run the main function
if __name__ == "__main__":
    search_and_process_data()
